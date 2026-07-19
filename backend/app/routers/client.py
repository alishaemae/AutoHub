
import os
import uuid as _uuid
from decimal import Decimal

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..database import get_db
from ..models import (CAR_STATUS_LABELS, Car, CarDocument, CarPhoto,
                     CurrencyRate, DepositReceipt, Role, Transaction,
                     TxDirection, User)
from ..schemas import (BalanceOut, CarOut, CurrencyOut, TransactionOut)
from ..security import get_current_user

router = APIRouter(prefix="/api", tags=["client"])

UPLOAD_DIR = "uploads"

_RECEIPT_EXT = (".pdf", ".png", ".jpg", ".jpeg", ".webp")


def _save_receipt(file: UploadFile, prefix: str) -> tuple[str, str]:
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in _RECEIPT_EXT:
        raise HTTPException(status_code=400,
                            detail="Чек: допустимы PDF или изображение")
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    path = os.path.join(UPLOAD_DIR, f"{prefix}_{_uuid.uuid4().hex[:8]}{ext}")
    return path, ext


def _car_to_out(car: Car) -> CarOut:
    out = CarOut.model_validate(car)
    out.status_label = CAR_STATUS_LABELS.get(car.status, "")
    out.has_photo = bool(car.photo_url)
    out.opis_count = len(car.photos)
    return out


@router.get("/cars", response_model=list[CarOut])
async def list_cars(user: User = Depends(get_current_user),
                    db: AsyncSession = Depends(get_db)):
    q = (select(Car).options(selectinload(Car.docs), selectinload(Car.photos))
         .order_by(Car.purchase_date.desc()))
    if user.role != Role.admin:
        q = q.where(Car.owner_id == user.id)
    res = await db.execute(q)
    return [_car_to_out(c) for c in res.scalars().all()]


@router.get("/cars/{car_id}", response_model=CarOut)
async def get_car(car_id: int, user: User = Depends(get_current_user),
                  db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(Car).options(selectinload(Car.docs), selectinload(Car.photos))
        .where(Car.id == car_id))
    car = res.scalar_one_or_none()
    if not car or (user.role != Role.admin and car.owner_id != user.id):
        raise HTTPException(status_code=404, detail="Автомобиль не найден")
    return _car_to_out(car)


@router.get("/balance", response_model=BalanceOut)
async def get_balance(user: User = Depends(get_current_user),
                      db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(Transaction).where(Transaction.user_id == user.id)
        .order_by(Transaction.op_date.desc()))
    txs = res.scalars().all()
    income = [t for t in txs if t.direction == TxDirection.income]
    expense = [t for t in txs if t.direction == TxDirection.expense]
    inc_total = sum((t.amount for t in income), Decimal(0))
    exp_total = sum((t.amount for t in expense), Decimal(0))
    return BalanceOut(
        balance=inc_total - exp_total,
        income_total=inc_total, expense_total=exp_total,
        income=[TransactionOut.model_validate(t) for t in income],
        expense=[TransactionOut.model_validate(t) for t in expense])


@router.get("/balance/export")
async def export_balance(user_id: int | None = None,
                         user: User = Depends(get_current_user),
                         db: AsyncSession = Depends(get_db)):

    target_id = user.id
    if user_id is not None and user.role == Role.admin:
        target_id = user_id
    owner = await db.get(User, target_id)
    res = await db.execute(
        select(Transaction).where(Transaction.user_id == target_id)
        .order_by(Transaction.op_date.desc()))
    txs = res.scalars().all()

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Выписка"

    RED = "C63E45"
    money_fmt = "#,##0.00 ₽"
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)


    ws.merge_cells("A1:E1")
    ws["A1"] = f"Выписка по балансу — {owner.full_name if owner else ''}".strip(" —")
    ws["A1"].font = Font(bold=True, size=14, color=RED)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 22


    headers = ["Дата", "Тип", "Наименование", "Плательщик", "Сумма"]
    hdr_row = 3
    for col, name in enumerate(headers, 1):
        c = ws.cell(row=hdr_row, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[hdr_row].height = 18

    inc = Decimal(0)
    exp = Decimal(0)
    r = hdr_row + 1
    for t in txs:
        income = t.direction == TxDirection.income
        ws.cell(row=r, column=1, value=t.op_date).number_format = "DD.MM.YYYY"
        ws.cell(row=r, column=2, value="Поступление" if income else "Списание")
        ws.cell(row=r, column=3, value=t.name)
        ws.cell(row=r, column=4, value=t.payer or "")
        amt = ws.cell(row=r, column=5, value=float(t.amount))
        amt.number_format = money_fmt
        amt.font = Font(color="2E7D46" if income else RED)
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = border
        if income:
            inc += t.amount
        else:
            exp += t.amount
        r += 1


    r += 1
    for label, val, color in [("Поступления", inc, "2E7D46"),
                              ("Списания", exp, RED),
                              ("Баланс", inc - exp, "1F2530")]:
        ws.cell(row=r, column=4, value=label).font = Font(bold=True)
        c = ws.cell(row=r, column=5, value=float(val))
        c.number_format = money_fmt
        c.font = Font(bold=True, color=color)
        r += 1

    widths = [13, 14, 40, 22, 16]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="balance-{target_id}.xlsx"'})


@router.get("/currency", response_model=list[CurrencyOut])
async def list_currency(user: User = Depends(get_current_user),
                        db: AsyncSession = Depends(get_db)):

    res = await db.execute(select(CurrencyRate).order_by(CurrencyRate.code))
    return res.scalars().all()


DEPOSIT_DEFAULT = ("Оплата депозита для доступа к ставкам на автомобили — 100 000 рублей.\n"
                   "В назначении платежа указать: Аванс по Агентскому договору №___ от ___")


@router.get("/deposit")
async def get_deposit(user: User = Depends(get_current_user),
                      db: AsyncSession = Depends(get_db)):
    """Текст и наличие QR для вкладки «Депозит»."""
    from ..models import Setting
    row = await db.get(Setting, "deposit_text")
    text = row.value if row and row.value else DEPOSIT_DEFAULT
    def _has(base):
        p = os.path.join(UPLOAD_DIR, base)
        return any(os.path.exists(p + ext) for ext in (".png", ".jpg", ".jpeg"))
    return {"text": text, "has_qr": _has("deposit_qr"),
            "has_qr_tbank": _has("deposit_qr_tbank")}


@router.get("/deposit/qr")
async def get_deposit_qr(bank: str = "sber", user: User = Depends(get_current_user)):
    base = "deposit_qr" if bank == "sber" else "deposit_qr_" + bank
    for ext in (".png", ".jpg", ".jpeg"):
        p = os.path.join(UPLOAD_DIR, base + ext)
        if os.path.exists(p):
            return FileResponse(p)
    raise HTTPException(status_code=404, detail="QR не загружен")


async def _owned_car_doc(doc_id: int, user: User, db: AsyncSession) -> CarDocument:
    res = await db.execute(select(CarDocument).options(selectinload(CarDocument.car))
                           .where(CarDocument.id == doc_id))
    doc = res.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Документ не найден")
    if user.role != Role.admin and doc.car.owner_id != user.id:
        raise HTTPException(status_code=404, detail="Документ не найден")
    return doc


@router.post("/car-documents/{doc_id}/receipt")
async def upload_car_receipt(doc_id: int, file: UploadFile = File(...),
                             user: User = Depends(get_current_user),
                             db: AsyncSession = Depends(get_db)):
    doc = await _owned_car_doc(doc_id, user, db)
    path, _ = _save_receipt(file, f"receipt_doc{doc_id}")
    if doc.receipt_path and os.path.exists(doc.receipt_path):
        try:
            os.remove(doc.receipt_path)
        except OSError:
            pass
    with open(path, "wb") as f:
        f.write(await file.read())
    doc.receipt_path = path
    doc.receipt_name = file.filename or os.path.basename(path)
    from datetime import datetime, timezone
    doc.receipt_uploaded_at = datetime.now(timezone.utc)
    await db.commit()
    return {"ok": True}


@router.get("/car-documents/{doc_id}/receipt")
async def download_car_receipt(doc_id: int, user: User = Depends(get_current_user),
                               db: AsyncSession = Depends(get_db)):
    doc = await _owned_car_doc(doc_id, user, db)
    if not doc.receipt_path or not os.path.exists(doc.receipt_path):
        raise HTTPException(status_code=404, detail="Чек не загружен")
    return FileResponse(doc.receipt_path, filename=doc.receipt_name or "чек")


@router.post("/deposit/receipt")
async def upload_deposit_receipt(file: UploadFile = File(...),
                                 user: User = Depends(get_current_user),
                                 db: AsyncSession = Depends(get_db)):
    path, _ = _save_receipt(file, f"deposit_receipt_u{user.id}")
    with open(path, "wb") as f:
        f.write(await file.read())
    rec = DepositReceipt(user_id=user.id,
                         filename=file.filename or os.path.basename(path),
                         file_path=path)
    db.add(rec)
    await db.commit()
    return {"ok": True}


@router.get("/deposit/receipts")
async def my_deposit_receipts(user: User = Depends(get_current_user),
                              db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(DepositReceipt)
                           .where(DepositReceipt.user_id == user.id)
                           .order_by(DepositReceipt.uploaded_at.desc()))
    return [{"id": r.id, "filename": r.filename,
             "uploaded_at": r.uploaded_at.isoformat() if r.uploaded_at else ""}
            for r in res.scalars().all()]


@router.get("/deposit/receipts/{rec_id}")
async def download_deposit_receipt(rec_id: int,
                                   user: User = Depends(get_current_user),
                                   db: AsyncSession = Depends(get_db)):
    rec = await db.get(DepositReceipt, rec_id)
    if not rec or (user.role != Role.admin and rec.user_id != user.id):
        raise HTTPException(status_code=404, detail="Чек не найден")
    if not os.path.exists(rec.file_path):
        raise HTTPException(status_code=404, detail="Файл отсутствует")
    return FileResponse(rec.file_path, filename=rec.filename or "чек")


@router.get("/documents/{doc_id}/download")
async def download_document(doc_id: int, user: User = Depends(get_current_user),
                            db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(CarDocument).options(selectinload(CarDocument.car))
        .where(CarDocument.id == doc_id))
    doc = res.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Документ не найден")

    if user.role != Role.admin and doc.car.owner_id != user.id:
        raise HTTPException(status_code=403, detail="Нет доступа")
    if not os.path.exists(doc.file_path):
        raise HTTPException(status_code=404, detail="Файл отсутствует")
    return FileResponse(doc.file_path, media_type="application/pdf",
                        filename=doc.filename)


async def _car_for_access(car_id: int, user: User, db: AsyncSession) -> Car:
    car = await db.get(Car, car_id)
    if not car or (user.role != Role.admin and car.owner_id != user.id):
        raise HTTPException(status_code=404, detail="Автомобиль не найден")
    return car


@router.get("/cars/{car_id}/photo")
async def car_photo(car_id: int, user: User = Depends(get_current_user),
                    db: AsyncSession = Depends(get_db)):
    car = await _car_for_access(car_id, user, db)
    if not car.photo_url or not os.path.exists(car.photo_url):
        raise HTTPException(status_code=404, detail="Фото отсутствует")
    return FileResponse(car.photo_url)


@router.get("/cars/{car_id}/opis")
async def car_opis_list(car_id: int, user: User = Depends(get_current_user),
                        db: AsyncSession = Depends(get_db)):
    """Список id фотографий фотоописи (для отрисовки сетки)."""
    await _car_for_access(car_id, user, db)
    res = await db.execute(
        select(CarPhoto).where(CarPhoto.car_id == car_id).order_by(CarPhoto.id))
    return [{"id": p.id} for p in res.scalars().all()]


@router.get("/cars/{car_id}/opis/{photo_id}")
async def car_opis_photo(car_id: int, photo_id: int,
                         user: User = Depends(get_current_user),
                         db: AsyncSession = Depends(get_db)):
    await _car_for_access(car_id, user, db)
    photo = await db.get(CarPhoto, photo_id)
    if not photo or photo.car_id != car_id or not os.path.exists(photo.file_path):
        raise HTTPException(status_code=404, detail="Фото отсутствует")
    return FileResponse(photo.file_path)


@router.get("/cars/{car_id}/archive")
async def car_archive(car_id: int, user: User = Depends(get_current_user),
                      db: AsyncSession = Depends(get_db)):
    """Собрать ZIP из фотографий фотоописи на лету и отдать на скачивание."""
    await _car_for_access(car_id, user, db)
    res = await db.execute(
        select(CarPhoto).where(CarPhoto.car_id == car_id).order_by(CarPhoto.id))
    photos = res.scalars().all()
    if not photos:
        raise HTTPException(status_code=404, detail="Фотоопись пуста")
    import io
    import zipfile
    from fastapi.responses import Response
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for i, p in enumerate(photos, 1):
            if os.path.exists(p.file_path):
                ext = os.path.splitext(p.file_path)[1] or ".jpg"
                z.write(p.file_path, f"photo_{i:02d}{ext}")
    buf.seek(0)
    return Response(
        content=buf.getvalue(), media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="photoopis-{car_id}.zip"'})
